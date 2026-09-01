"""
api/export.py — Advanced Export Service (P9).

Provides:
- PDF export with ReportLab
- Excel export with openpyxl
- CSV export
- JSON export
- Custom report templates & export history tracking
- ResultStore (P5) integration for secure result_id reference

Exposes endpoints under ``/api/v1/export``:
* ``POST /{project_id}/pdf``     — Export study results as PDF
* ``POST /{project_id}/excel``   — Export study results as Excel
* ``POST /{project_id}/csv``     — Export study results as CSV
* ``POST /{project_id}/json``    — Export study results as JSON
* ``GET  /history``              — List export history
* ``GET  /formats``              — List pre-declared supported export formats
"""

from __future__ import annotations

import csv
import io
import json
import logging
import re
import uuid
from datetime import datetime, timezone
from typing import Any, Optional, Sequence

from fastapi import APIRouter, Depends, Header, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, ConfigDict
from sqlalchemy import (
    DateTime,
    String,
    desc,
    func,
    select,
)
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import Mapped, mapped_column

from api._messages import MSG_PROJECT_NOT_FOUND
from api.database import Base, get_db
from api.dependencies import (
    CurrentUser,
    PaginationParams,
    get_api_key,
    pagination_params,
)
from api.dual_control import record_approval_event
from api.feature_flags import is_feature_enabled
from api.rbac import require_permission
from api.results_store import (
    create_result,
    store_result_file,
)

logger = logging.getLogger("api.export")
UTC = timezone.utc


def _sanitize_filename(name: str, max_length: int = 64) -> str:
    """Sanitize a string for use in a Content-Disposition filename."""
    if not name:
        return "untitled"
    sanitized = re.sub(r'[\r\n"\x00-\x1f\x7f/\\]', "", str(name))
    sanitized = re.sub(r"\s+", "_", sanitized).strip("._")
    if not sanitized:
        sanitized = "untitled"
    return sanitized[:max_length]


class ExportHistory(Base):
    """Track export operations in database."""

    __tablename__ = "export_history"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    project_id: Mapped[str] = mapped_column(String(36), index=True, nullable=False)
    study_id: Mapped[Optional[str]] = mapped_column(String(36), nullable=True)
    export_type: Mapped[str] = mapped_column(String(16), nullable=False)  # pdf, excel, csv, json
    file_name: Mapped[str] = mapped_column(String(255), nullable=False)
    file_size_bytes: Mapped[Optional[int]] = mapped_column(nullable=True)
    created_by: Mapped[str] = mapped_column(String(36), nullable=False)
    created_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        default=lambda: datetime.now(UTC),
    )


class ExportFormat(BaseModel):
    """A supported export format."""

    id: str
    name: str
    mime_type: str
    extension: str
    description: str


class ExportResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: str
    project_id: str
    study_id: str | None = None
    export_type: str
    file_name: str
    file_size_bytes: int | None = None
    result_id: str | None = None
    created_by: str = ""
    created_at: datetime | None = None


class ExportHistoryResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    exports: list[ExportResponse]
    total: int


EXPORT_FORMATS: list[ExportFormat] = [
    ExportFormat(
        id="pdf",
        name="PDF Report",
        mime_type="application/pdf",
        extension=".pdf",
        description="Comprehensive formatted engineering report with tables and diagrams",
    ),
    ExportFormat(
        id="excel",
        name="Excel Workbook",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        extension=".xlsx",
        description="Structured study results and bus/branch matrices in XLSX format",
    ),
    ExportFormat(
        id="csv",
        name="CSV Data",
        mime_type="text/csv",
        extension=".csv",
        description="Raw bus and branch table exports in comma-separated values format",
    ),
    ExportFormat(
        id="json",
        name="JSON Data",
        mime_type="application/json",
        extension=".json",
        description="Full power system model and study results in JSON format",
    ),
]

router = APIRouter(
    prefix="/api/v1/export", tags=["Export"], dependencies=[Depends(get_api_key)]
)


async def _get_project_studies(project_id: str, db: AsyncSession) -> Sequence[Any]:
    from api.projects import StudyResult

    result = await db.execute(
        select(StudyResult)
        .where(StudyResult.project_id == project_id)
        .order_by(desc(StudyResult.created_at))
    )
    return list(result.scalars().all())


async def _load_owned_project(project_id: str, user: CurrentUser, db: AsyncSession):
    """Load a project owned by user (or admin), returning 404 on any mismatch."""
    from api.projects import Project

    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail=MSG_PROJECT_NOT_FOUND)
    if user.role != "admin" and project.created_by != user.user_id:
        raise HTTPException(status_code=404, detail=MSG_PROJECT_NOT_FOUND)
    return project


def _generate_pdf(project_name: str, studies: Sequence[Any]) -> bytes:
    """Generate a PDF report using ReportLab."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError:
        content = f"PDF Export - {project_name}\n\n"
        for s in studies:
            content += f"Study: {s.study_type} - Status: {s.status}\n"
        return content.encode("utf-8")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements: list[Any] = []

    elements.append(Paragraph(f"Project Report: {project_name}", styles["Title"]))
    elements.append(Spacer(1, 12))
    elements.append(
        Paragraph(
            f"Generated: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]
        )
    )
    elements.append(Spacer(1, 24))

    data = [["Study Type", "Status", "Created", "Results"]]
    for s in studies:
        results_summary = ""
        if s.results:
            results_summary = ", ".join(list(s.results.keys())[:3])
        data.append(
            [
                s.study_type,
                s.status,
                s.created_at.strftime("%Y-%m-%d") if s.created_at else "",
                results_summary,
            ]
        )

    table = Table(data, colWidths=[120, 80, 100, 200])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]
        )
    )
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def _generate_excel(project_name: str, studies: Sequence[Any]) -> bytes:
    """Generate an Excel file using openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        content = "Study Type,Status,Created,Results\n"
        for s in studies:
            results_str = json.dumps(s.results) if s.results else ""
            content += f"{s.study_type},{s.status},{s.created_at},{results_str}\n"
        return content.encode("utf-8")

    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet(title="Study Results")
    else:
        ws.title = "Study Results"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    headers = ["Study Type", "Status", "Created At", "Results Summary"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row, s in enumerate(studies, 2):
        ws.cell(row=row, column=1, value=s.study_type)
        ws.cell(row=row, column=2, value=s.status)
        ws.cell(row=row, column=3, value=str(s.created_at) if s.created_at else "")
        ws.cell(row=row, column=4, value=json.dumps(s.results) if s.results else "")

    for col in range(1, 5):
        ws.column_dimensions[chr(64 + col)].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _generate_csv(project_name: str, studies: Sequence[Any]) -> bytes:
    """Generate CSV format study results."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["project_name", "study_type", "status", "created_at", "results"])
    for s in studies:
        results_str = json.dumps(s.results) if s.results else ""
        writer.writerow([project_name, s.study_type, s.status, str(s.created_at) if s.created_at else "", results_str])
    return output.getvalue().encode("utf-8")


def _generate_json(project_name: str, studies: Sequence[Any]) -> bytes:
    """Generate JSON format study results."""
    data = {
        "project_name": project_name,
        "exported_at": datetime.now(UTC).isoformat(),
        "studies": [
            {
                "study_type": s.study_type,
                "status": s.status,
                "created_at": s.created_at.isoformat() if s.created_at else None,
                "results": s.results,
            }
            for s in studies
        ],
    }
    return json.dumps(data, indent=2).encode("utf-8")


@router.get("/formats", summary="List pre-declared supported export formats")
async def list_export_formats() -> list[ExportFormat]:
    """Return all available export formats."""
    return EXPORT_FORMATS


@router.post("/{project_id}/pdf", responses={404: {"description": MSG_PROJECT_NOT_FOUND}})
async def export_pdf(
    project_id: str,
    db: AsyncSession = Depends(get_db),
    auth=Depends(require_permission("export", "create")),
    idempotency_key: Optional[str] = Header(default=None, alias="Idempotency-Key"),
):
    """Export study results as PDF."""
    if not is_feature_enabled("data_export", default=True):
        raise HTTPException(status_code=403, detail="Data export feature is disabled")

    user: CurrentUser = auth[0] if isinstance(auth, tuple) else auth
    project = await _load_owned_project(project_id, user, db)
    studies = await _get_project_studies(project_id, db)
    pdf_bytes = _generate_pdf(project.name, studies)

    safe_name = _sanitize_filename(project.name)
    file_name = f"{safe_name}_report.pdf"

    # Store in ResultStore
    result_id = None
    try:
        result_id = await create_result(
            tenant_id=user.tenant_id or "default",
            project_id=project_id,
            created_by=user.user_id,
            summary_json={
                "export_type": "pdf",
                "project_name": project.name,
                "file_name": file_name,
                "file_size_bytes": len(pdf_bytes),
            },
            ttl_days=30,
        )
        await store_result_file(
            tenant_id=user.tenant_id or "default",
            result_id=result_id,
            rel_path=file_name,
            data=pdf_bytes,
            mime="application/pdf",
        )
    except Exception:
        logger.debug("Failed to store export in ResultStore", exc_info=True)

    export = ExportHistory(
        id=str(uuid.uuid4()),
        project_id=project_id,
        export_type="pdf",
        file_name=file_name,
        file_size_bytes=len(pdf_bytes),
        created_by=user.user_id,
    )
    db.add(export)
    await db.flush()

    record_approval_event(
        "EXPORT_REQUESTED",
        export.id,
        user.user_id,
        {"project_id": project_id, "export_type": "pdf", "file_name": file_name, "result_id": result_id},
    )

    headers = {
        "Content-Disposition": f'attachment; filename="{file_name}"',
    }
    if result_id:
        headers["X-Result-ID"] = result_id

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers=headers,
    )


@router.post("/{project_id}/excel", responses={404: {"description": MSG_PROJECT_NOT_FOUND}})
async def export_excel(
    project_id: str,
    db: AsyncSession = Depends(get_db),
    auth=Depends(require_permission("export", "create")),
    idempotency_key: Optional[str] = Header(default=None, alias="Idempotency-Key"),
):
    """Export study results as Excel."""
    if not is_feature_enabled("data_export", default=True):
        raise HTTPException(status_code=403, detail="Data export feature is disabled")

    user: CurrentUser = auth[0] if isinstance(auth, tuple) else auth
    project = await _load_owned_project(project_id, user, db)
    studies = await _get_project_studies(project_id, db)
    excel_bytes = _generate_excel(project.name, studies)

    safe_name = _sanitize_filename(project.name)
    file_name = f"{safe_name}_results.xlsx"

    result_id = None
    try:
        result_id = await create_result(
            tenant_id=user.tenant_id or "default",
            project_id=project_id,
            created_by=user.user_id,
            summary_json={
                "export_type": "excel",
                "project_name": project.name,
                "file_name": file_name,
                "file_size_bytes": len(excel_bytes),
            },
            ttl_days=30,
        )
        await store_result_file(
            tenant_id=user.tenant_id or "default",
            result_id=result_id,
            rel_path=file_name,
            data=excel_bytes,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception:
        logger.debug("Failed to store excel export in ResultStore", exc_info=True)

    export = ExportHistory(
        id=str(uuid.uuid4()),
        project_id=project_id,
        export_type="excel",
        file_name=file_name,
        file_size_bytes=len(excel_bytes),
        created_by=user.user_id,
    )
    db.add(export)
    await db.flush()

    record_approval_event(
        "EXPORT_REQUESTED",
        export.id,
        user.user_id,
        {"project_id": project_id, "export_type": "excel", "file_name": file_name, "result_id": result_id},
    )

    headers = {
        "Content-Disposition": f'attachment; filename="{file_name}"',
    }
    if result_id:
        headers["X-Result-ID"] = result_id

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/{project_id}/csv", responses={404: {"description": MSG_PROJECT_NOT_FOUND}})
async def export_csv(
    project_id: str,
    db: AsyncSession = Depends(get_db),
    auth=Depends(require_permission("export", "create")),
):
    """Export study results as CSV."""
    if not is_feature_enabled("data_export", default=True):
        raise HTTPException(status_code=403, detail="Data export feature is disabled")

    user: CurrentUser = auth[0] if isinstance(auth, tuple) else auth
    project = await _load_owned_project(project_id, user, db)
    studies = await _get_project_studies(project_id, db)
    csv_bytes = _generate_csv(project.name, studies)

    safe_name = _sanitize_filename(project.name)
    file_name = f"{safe_name}_results.csv"

    export = ExportHistory(
        id=str(uuid.uuid4()),
        project_id=project_id,
        export_type="csv",
        file_name=file_name,
        file_size_bytes=len(csv_bytes),
        created_by=user.user_id,
    )
    db.add(export)
    await db.flush()

    record_approval_event(
        "EXPORT_REQUESTED",
        export.id,
        user.user_id,
        {"project_id": project_id, "export_type": "csv", "file_name": file_name},
    )

    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
    )


@router.post("/{project_id}/json", responses={404: {"description": MSG_PROJECT_NOT_FOUND}})
async def export_json(
    project_id: str,
    db: AsyncSession = Depends(get_db),
    auth=Depends(require_permission("export", "create")),
):
    """Export study results as JSON."""
    if not is_feature_enabled("data_export", default=True):
        raise HTTPException(status_code=403, detail="Data export feature is disabled")

    user: CurrentUser = auth[0] if isinstance(auth, tuple) else auth
    project = await _load_owned_project(project_id, user, db)
    studies = await _get_project_studies(project_id, db)
    json_bytes = _generate_json(project.name, studies)

    safe_name = _sanitize_filename(project.name)
    file_name = f"{safe_name}_results.json"

    export = ExportHistory(
        id=str(uuid.uuid4()),
        project_id=project_id,
        export_type="json",
        file_name=file_name,
        file_size_bytes=len(json_bytes),
        created_by=user.user_id,
    )
    db.add(export)
    await db.flush()

    record_approval_event(
        "EXPORT_REQUESTED",
        export.id,
        user.user_id,
        {"project_id": project_id, "export_type": "json", "file_name": file_name},
    )

    return StreamingResponse(
        io.BytesIO(json_bytes),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
    )


@router.get("/history", response_model=ExportHistoryResponse)
async def export_history(
    db: AsyncSession = Depends(get_db),
    auth=Depends(require_permission("export", "list")),
    pagination: PaginationParams = Depends(pagination_params),
):
    """List export history scoped to the authenticated user."""
    user: CurrentUser = auth[0] if isinstance(auth, tuple) else auth
    result = await db.execute(
        select(ExportHistory)
        .where(ExportHistory.created_by == user.user_id)
        .order_by(desc(ExportHistory.created_at))
        .offset(pagination.offset)
        .limit(pagination.page_size)
    )
    exports = result.scalars().all()
    count = await db.execute(
        select(func.count())
        .select_from(ExportHistory)
        .where(ExportHistory.created_by == user.user_id)
    )
    total = count.scalar_one()
    return ExportHistoryResponse(
        exports=[
            ExportResponse(
                id=e.id,
                project_id=e.project_id,
                study_id=e.study_id,
                export_type=e.export_type,
                file_name=e.file_name,
                file_size_bytes=e.file_size_bytes,
                created_by=e.created_by,
                created_at=e.created_at,
            )
            for e in exports
        ],
        total=total,
    )
